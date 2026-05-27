# ============================================================
# Python + Requests 14天学习计划
# Day 10 - 数据驱动测试（CSV/Excel读取 + parametrize进阶）
# ============================================================
"""
今日学习目标：
1. 理解数据驱动测试（DDT）的核心思想
2. 用 CSV 文件管理测试数据 + pytest.mark.parametrize 驱动用例
3. 用 openpyxl 读取 Excel 测试数据
4. parametrize 进阶：ids 命名 / 间接参数化 / 多参数组合
5. 项目实战：把山海之巅接口测试数据全部抽离到数据文件

依赖安装（如未安装）：
    pip install openpyxl pandas
"""

import csv
import json
import os
import time
import requests
import pytest

# ============================================================
# Part 0：前置 - 公共 HTTP 工具
# ============================================================
BASE_URL = "https://httpbin.org"

def safe_get(data, *keys, default=None):
    """安全多层取值，防 KeyError / TypeError"""
    try:
        for key in keys:
            data = data[key]
        return data
    except (KeyError, TypeError, IndexError):
        return default


def http_request(method, path, **kwargs):
    """统一 HTTP 请求封装，自动重试3次应对 httpbin.org 偶发抖动"""
    url = BASE_URL + path
    max_retry = 3
    for attempt in range(1, max_retry + 1):
        try:
            resp = requests.request(method, url, timeout=15, **kwargs)
            if resp.status_code < 500:
                return resp
            print(f"  [retry {attempt}/{max_retry}] {resp.status_code} - 等待重试...")
        except requests.exceptions.RequestException as e:
            print(f"  [retry {attempt}/{max_retry}] 请求异常: {e}")
        if attempt < max_retry:
            time.sleep(1.5)
    raise RuntimeError(f"请求 {method} {url} 连续失败 {max_retry} 次")


# ============================================================
# Part 1：数据驱动测试核心思想
# ============================================================
print("=" * 60)
print("【Part 1】数据驱动测试（DDT）核心思想")
print("=" * 60)

"""
传统写法（数据和代码混在一起）：

def test_login_1():
    r = requests.post(..., json={"user": "admin", "pass": "123"})
    assert r.status_code == 200

def test_login_2():
    r = requests.post(..., json={"user": "", "pass": ""})
    assert r.status_code == 400

-- 缺点：100个场景 = 100个函数，维护噩梦 --

DDT写法（数据和代码分离）：
- 测试逻辑写一次
- 测试数据放到 CSV / Excel / JSON 文件
- parametrize 自动展开为多个用例
"""

ddt_demo_data = [
    {"scene": "正常登录",   "username": "admin",  "password": "123456",  "expected": 200},
    {"scene": "空用户名",   "username": "",       "password": "123456",  "expected": 400},
    {"scene": "空密码",    "username": "admin",  "password": "",        "expected": 400},
    {"scene": "SQL注入",   "username": "' OR '1", "password": "x",      "expected": 400},
]

print("\n数据驱动示意（数据来自列表）：")
for item in ddt_demo_data:
    print(f"  场景={item['scene']:<8} 预期状态码={item['expected']}")


# ============================================================
# Part 2：CSV 文件读写 + 测试数据管理
# ============================================================
print("\n" + "=" * 60)
print("【Part 2】CSV 文件读写 + 测试数据管理")
print("=" * 60)

# 2-1 生成测试数据 CSV
CSV_PATH = os.path.join(os.path.dirname(__file__), "test_data_login.csv")

login_test_data = [
    ["scene", "method", "path", "username", "password", "expected_status", "expected_key"],
    ["正常GET请求",   "GET",  "/get",  "",      "",       "200", "url"],
    ["带参数GET",    "GET",  "/get",  "admin", "dummy",  "200", "args"],
    ["正常POST请求",  "POST", "/post", "admin", "123456", "200", "json"],
    ["空Body POST",  "POST", "/post", "",      "",       "200", "data"],
    ["DELETE请求",   "DELETE", "/delete", "", "",        "200", "url"],
]

with open(CSV_PATH, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)
    writer.writerows(login_test_data)

print(f"\n已生成测试数据 CSV：{CSV_PATH}")

# 2-2 读取 CSV 并展示
print("\n读取 CSV 内容：")
with open(CSV_PATH, "r", encoding="utf-8") as f:
    reader = csv.DictReader(f)
    rows = list(reader)
    for i, row in enumerate(rows, 1):
        print(f"  [{i}] {row['scene']:<12} {row['method']:<8} {row['path']}")

print(f"\n共 {len(rows)} 条测试数据")


# 2-3 工具函数：从 CSV 加载测试数据
def load_csv_data(filepath):
    """从 CSV 文件加载测试数据，返回 list[dict]"""
    with open(filepath, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return list(reader)


# ============================================================
# Part 3：Excel 文件读写（openpyxl）
# ============================================================
print("\n" + "=" * 60)
print("【Part 3】Excel 文件读写（openpyxl）")
print("=" * 60)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "test_data_api.xlsx")

try:
    import openpyxl

    # 3-1 创建测试数据 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "登录接口"

    headers = ["用例编号", "用例名称", "请求方法", "接口路径", "请求参数",
               "预期状态码", "预期字段", "备注"]
    ws.append(headers)

    api_cases = [
        ["TC001", "正常GET请求",   "GET",    "/get",     "",                          200, "url",     "基础冒烟"],
        ["TC002", "带查询参数GET", "GET",    "/get",     '{"name":"test","page":1}',  200, "args",    "参数传递"],
        ["TC003", "JSON_POST",    "POST",   "/post",    '{"user":"admin","pw":"123"}',200, "json",    "JSON Body"],
        ["TC004", "表单POST",     "POST",   "/post",    '{"key":"val"}',             200, "form",    "Form数据"],
        ["TC005", "PUT更新",      "PUT",    "/put",     '{"id":1,"name":"updated"}', 200, "json",    "更新资源"],
        ["TC006", "DELETE删除",   "DELETE", "/delete",  "",                          200, "url",     "删除资源"],
        ["TC007", "PATCH部分更新","PATCH",  "/patch",   '{"status":"active"}',       200, "json",    "部分更新"],
        ["TC008", "自定义Headers","GET",    "/headers", "",                          200, "headers", "请求头检查"],
    ]
    for case in api_cases:
        ws.append(case)

    # 调整列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    wb.save(EXCEL_PATH)
    print(f"\n已生成测试数据 Excel：{EXCEL_PATH}")

    # 3-2 读取 Excel
    wb_read = openpyxl.load_workbook(EXCEL_PATH)
    ws_read = wb_read.active
    excel_rows = []
    headers_row = [cell.value for cell in ws_read[1]]
    print(f"\n表头：{headers_row}")
    for row in ws_read.iter_rows(min_row=2, values_only=True):
        if row[0]:  # 跳过空行
            excel_rows.append(dict(zip(headers_row, row)))

    print(f"\n读取 Excel 内容（共 {len(excel_rows)} 条）：")
    for case in excel_rows:
        print(f"  {case['用例编号']}: {case['用例名称']:<12} {case['请求方法']:<8} {case['接口路径']}")

    # 3-3 工具函数
    def load_excel_data(filepath, sheet_name=None):
        """从 Excel 读取测试数据，返回 list[dict]"""
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name] if sheet_name else wb.active
        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                data.append(dict(zip(headers, row)))
        return data

    EXCEL_AVAILABLE = True
    print("\n✅ openpyxl 加载成功，Excel 测试数据就绪")

except ImportError:
    print("\n⚠️  openpyxl 未安装，跳过 Excel 部分")
    print("    安装命令：pip install openpyxl")
    EXCEL_AVAILABLE = False


# ============================================================
# Part 4：parametrize 进阶用法
# ============================================================
print("\n" + "=" * 60)
print("【Part 4】parametrize 进阶用法详解")
print("=" * 60)

print("""
一、基础用法（单参数）
    @pytest.mark.parametrize("status_code", [200, 201, 204])
    def test_status(status_code):
        assert status_code < 300

二、多参数解包（元组）
    @pytest.mark.parametrize("method,path,expected", [
        ("GET",  "/get",    200),
        ("POST", "/post",   200),
    ])
    def test_api(method, path, expected):
        ...

三、ids 命名（让报告更可读）
    @pytest.mark.parametrize("x,y", [(1,2),(3,4)], ids=["case1","case2"])

四、从文件加载数据（数据驱动核心）
    CSV_DATA = load_csv_data("test_data.csv")
    
    @pytest.mark.parametrize("row", CSV_DATA, ids=[r["scene"] for r in CSV_DATA])
    def test_from_csv(row):
        resp = requests.request(row["method"], BASE_URL + row["path"])
        assert resp.status_code == int(row["expected_status"])

五、indirect 间接参数化（配合 fixture）
    @pytest.mark.parametrize("token", ["valid_token", "expired_token"],
                              indirect=True)
    def test_with_fixture(token):
        ...  # token 会先经过同名 fixture 处理
""")


# ============================================================
# Part 5：pytest 测试用例（数据驱动实战）
# ============================================================

# 5-1 从 CSV 加载数据
_csv_data = load_csv_data(CSV_PATH)

# 5-2 从 Excel 加载数据（如果可用）
_excel_data = []
if EXCEL_AVAILABLE:
    _excel_data = load_excel_data(EXCEL_PATH)

# ----- 测试用例区域 -----

class TestCSVDriven:
    """
    Part 5 - CSV 数据驱动测试
    直接从 CSV 文件读取参数，parametrize 自动展开为多个用例
    """

    @pytest.mark.parametrize(
        "row",
        _csv_data,
        ids=[r["scene"] for r in _csv_data]
    )
    def test_api_from_csv(self, row):
        """从 CSV 驱动 API 测试"""
        method = row["method"]
        path   = row["path"]
        expected_status = int(row["expected_status"])
        expected_key    = row["expected_key"]

        # 构造请求参数
        kwargs = {}
        if method == "GET" and row.get("username"):
            kwargs["params"] = {"username": row["username"], "password": row["password"]}
        elif method in ("POST", "PUT", "PATCH", "DELETE"):
            if row.get("username"):
                kwargs["json"] = {"username": row["username"], "password": row["password"]}

        resp = http_request(method, path, **kwargs)

        # 断言1：状态码
        assert resp.status_code == expected_status, \
            f"[{row['scene']}] 状态码不符: 期望{expected_status}，实际{resp.status_code}"

        # 断言2：响应包含预期字段
        if expected_key:
            body = resp.json()
            assert expected_key in body, \
                f"[{row['scene']}] 响应体缺少字段 '{expected_key}': {list(body.keys())}"

        print(f"  ✅ {row['scene']}: {method} {path} → {resp.status_code}")


class TestExcelDriven:
    """
    Part 5 - Excel 数据驱动测试
    从 xlsx 读取用例，支持中文表头
    """

    @pytest.mark.skipif(not EXCEL_AVAILABLE, reason="openpyxl 未安装")
    @pytest.mark.parametrize(
        "case",
        _excel_data,
        ids=[c["用例编号"] for c in _excel_data] if _excel_data else []
    )
    def test_api_from_excel(self, case):
        """从 Excel 驱动 API 测试"""
        method       = case["请求方法"]
        path         = case["接口路径"]
        expected_st  = int(case["预期状态码"])
        expected_key = case["预期字段"]
        raw_params   = case.get("请求参数") or ""

        # 解析请求参数（JSON字符串→dict）
        kwargs = {}
        if raw_params.strip():
            try:
                params_dict = json.loads(raw_params)
                if method == "GET":
                    kwargs["params"] = params_dict
                else:
                    kwargs["json"] = params_dict
            except json.JSONDecodeError:
                pass  # 参数格式异常不影响主流程

        resp = http_request(method, path, **kwargs)

        # 断言1：状态码
        assert resp.status_code == expected_st, \
            f"[{case['用例名称']}] 状态码不符: 期望{expected_st}，实际{resp.status_code}"

        # 断言2：响应包含预期字段
        if expected_key:
            body = resp.json()
            assert expected_key in body, \
                f"[{case['用例名称']}] 响应体缺少字段 '{expected_key}'"

        print(f"  ✅ {case['用例编号']} {case['用例名称']}: {method} {path} → {resp.status_code}")


class TestParametrizeAdvanced:
    """
    Part 6 - parametrize 进阶技巧
    """

    # 6-1 多参数解包（元组列表）
    @pytest.mark.parametrize(
        "method, path, expected_key",
        [
            ("GET",    "/get",    "url"),
            ("POST",   "/post",   "json"),
            ("PUT",    "/put",    "json"),
            ("DELETE", "/delete", "url"),
            ("PATCH",  "/patch",  "json"),
        ],
        ids=["GET", "POST", "PUT", "DELETE", "PATCH"]
    )
    def test_multi_param(self, method, path, expected_key):
        """多参数解包 parametrize"""
        payload = {"test": "day10"} if method not in ("GET", "DELETE") else None
        resp = http_request(method, path, json=payload)

        assert resp.status_code == 200, f"{method} {path} 失败: {resp.status_code}"
        body = resp.json()
        assert expected_key in body, f"缺少字段 '{expected_key}': {list(body.keys())}"
        print(f"  ✅ {method} {path} → 包含字段 '{expected_key}'")

    # 6-2 用 ids 参数让测试报告更可读
    @pytest.mark.parametrize(
        "status_path, expected",
        [
            ("/status/200", 200),
            ("/status/201", 201),
            ("/status/204", 204),
            ("/status/400", 400),
            ("/status/404", 404),
        ],
        ids=["正常200", "创建201", "无内容204", "错误400", "未找到404"]
    )
    def test_status_codes_with_ids(self, status_path, expected):
        """验证各种 HTTP 状态码（ids让报告显示中文名称）"""
        resp = http_request("GET", status_path)
        assert resp.status_code == expected, \
            f"路径 {status_path}: 期望{expected}，实际{resp.status_code}"
        print(f"  ✅ {status_path} → {resp.status_code}")

    # 6-3 内联数据 + ids 组合（山海之巅模拟场景）
    @pytest.mark.parametrize(
        "scene, payload, check_key",
        [
            ("普通玩家登录",  {"username": "player1",  "password": "pass123"},   "json"),
            ("VIP玩家登录",  {"username": "vip_user",  "password": "vippass"},   "json"),
            ("GM账号登录",   {"username": "gm_admin",  "password": "gm123456"},  "json"),
            ("游客登录",    {"username": "guest_001", "password": ""},          "json"),
        ],
        ids=lambda x: x if isinstance(x, str) else None
    )
    def test_shanhai_login_scenarios(self, scene, payload, check_key):
        """山海之巅登录场景矩阵测试（数据内联）"""
        resp = http_request("POST", "/post", json=payload)

        assert resp.status_code == 200, f"[{scene}] 登录请求失败: {resp.status_code}"
        body = resp.json()
        assert check_key in body, f"[{scene}] 响应缺少字段 '{check_key}'"
        # 验证发送的数据被正确接收
        sent_json = safe_get(body, "json") or {}
        assert sent_json.get("username") == payload["username"], \
            f"[{scene}] username 传递不正确"
        print(f"  ✅ {scene}: username={payload['username']} 验证通过")


class TestBoundaryConditions:
    """
    Part 7 - 边界情况处理（数据驱动视角）
    """

    # 边界值测试数据
    BOUNDARY_DATA = [
        ("空字符串",       "",          True),
        ("单字符",        "a",         True),
        ("超长字符串",     "x" * 100,   True),
        ("特殊字符",       "!@#$%^&*()", True),
        ("中文字符",       "管理员",     True),
        ("SQL注入尝试",    "' OR '1'='1", True),
        ("换行符",        "line1\nline2", True),
    ]

    @pytest.mark.parametrize(
        "scene, value, should_accept",
        BOUNDARY_DATA,
        ids=[d[0] for d in BOUNDARY_DATA]
    )
    def test_boundary_username(self, scene, value, should_accept):
        """用户名边界值测试：验证接口能正确接收各类边界输入"""
        payload = {"username": value, "action": "login"}
        resp = http_request("POST", "/post", json=payload)

        # httpbin.org 接受所有输入，验证传输正确性
        assert resp.status_code == 200, f"[{scene}] 请求失败: {resp.status_code}"
        body = resp.json()
        received_username = safe_get(body, "json", "username", default="")

        if should_accept:
            assert received_username == value, \
                f"[{scene}] 值传递失败: 发送={repr(value)}, 接收={repr(received_username)}"
        print(f"  ✅ {scene}: 边界值传递 {'正确' if received_username == value else '异常'}")


# ============================================================
# Part 8：运行说明 & 总结
# ============================================================
if __name__ == "__main__":
    print("\n" + "=" * 60)
    print("【Day 10 总结】数据驱动测试核心要点")
    print("=" * 60)

    summary = """
✅ 掌握内容：
   1. DDT 思想：数据与逻辑分离，一份代码跑 N 个场景
   2. CSV 读写：csv.DictReader 读取，DictWriter 写入
   3. Excel 读写：openpyxl 读写，中文表头处理
   4. parametrize 进阶：
      - 基础单参数 / 多参数元组解包
      - ids 命名（中文名让报告更清晰）
      - 从文件加载数据（load_csv_data / load_excel_data）
      - lambda ids 过滤（跳过非字符串参数）
   5. 边界值矩阵：把边界数据集中到列表，参数化驱动

📂 生成的数据文件：
   - test_data_login.csv   （CSV 测试数据）
   - test_data_api.xlsx    （Excel 测试数据，需 openpyxl）

▶️  运行命令：
   pytest requests_day10.py -v -s
   pytest requests_day10.py -v -s -k "CSV"          # 只跑CSV驱动
   pytest requests_day10.py -v -s -k "Excel"         # 只跑Excel驱动
   pytest requests_day10.py -v -s -k "Advanced"      # 只跑进阶用法
   pytest requests_day10.py -v -s -k "Boundary"      # 只跑边界条件

📈 明日预告：Day 11 - 项目实战第一天
   用山海之巅接口文档编写真实测试套件，数据放 Excel 管理
    """
    print(summary)

    print("\n▶️  正在运行 pytest ...")
    import subprocess, sys
    result = subprocess.run(
        [sys.executable, "-m", "pytest", __file__, "-v", "-s", "--tb=short"],
        capture_output=False
    )
    sys.exit(result.returncode)
